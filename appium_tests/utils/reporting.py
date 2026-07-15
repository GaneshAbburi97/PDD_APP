from __future__ import annotations

import json
import re
from collections import Counter
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


STATUS_PASSED = "PASSED"
STATUS_FAILED = "FAILED"
STATUS_SKIPPED = "SKIPPED"


@dataclass
class TestResult:
    nodeid: str
    suite: str
    test_name: str
    markers: str
    status: str
    duration_seconds: float
    stage: str
    error: str
    screenshot: str
    timestamp: str


class TestResultCollector:
    def __init__(self, report_dir: Path):
        self.report_dir = report_dir
        self.screenshot_dir = report_dir / "screenshots"
        self.results: list[TestResult] = []

    def save_failure_screenshot(self, driver: Any, nodeid: str) -> Path | None:
        safe_name = re.sub(r"[^A-Za-z0-9_.-]+", "_", nodeid).strip("_")[:140]
        path = self.screenshot_dir / f"{datetime.now():%Y%m%d_%H%M%S}_{safe_name}.png"
        try:
            driver.save_screenshot(str(path))
            return path
        except Exception:
            return None

    def add_pytest_report(self, item: Any, report: Any, *, stage: str, screenshot_path: Path | None) -> None:
        if report.passed:
            status = STATUS_PASSED
        elif report.skipped:
            status = STATUS_SKIPPED
        else:
            status = STATUS_FAILED

        marker_names = sorted({marker.name for marker in item.iter_markers()})
        self.results.append(
            TestResult(
                nodeid=item.nodeid,
                suite=Path(str(item.fspath)).name,
                test_name=item.name,
                markers=", ".join(marker_names),
                status=status,
                duration_seconds=round(float(getattr(report, "duration", 0.0)), 3),
                stage=stage,
                error=self._format_error(report),
                screenshot=str(screenshot_path.resolve()) if screenshot_path else "",
                timestamp=datetime.now().isoformat(timespec="seconds"),
            )
        )

    def flush(self) -> dict[str, Path]:
        self.report_dir.mkdir(parents=True, exist_ok=True)
        self.screenshot_dir.mkdir(parents=True, exist_ok=True)
        json_path = self.report_dir / f"tmd_appium_results_{datetime.now():%Y%m%d_%H%M%S}.json"
        json_path.write_text(
            json.dumps([asdict(result) for result in self.results], indent=2),
            encoding="utf-8",
        )
        return {"json": json_path}

    @staticmethod
    def _format_error(report: Any) -> str:
        if not (report.failed or report.skipped):
            return ""
        longrepr = getattr(report, "longrepr", "")
        text = str(longrepr)
        return text[-32000:]


class ExcelReportBuilder:
    def __init__(
        self,
        *,
        results: list[TestResult],
        report_dir: Path,
        run_started_at: datetime,
        appium_config: Any,
        raw_results_path: Path,
    ):
        self.results = results
        self.report_dir = report_dir
        self.run_started_at = run_started_at
        self.run_finished_at = datetime.now()
        self.appium_config = appium_config
        self.raw_results_path = raw_results_path

    def build(self) -> Path:
        workbook = Workbook()
        summary = workbook.active
        summary.title = "Summary"
        details = workbook.create_sheet("Details")
        coverage = workbook.create_sheet("Coverage")
        actions = workbook.create_sheet("Recommendations")

        self._write_details(details)
        self._write_summary(summary)
        self._write_coverage(coverage)
        self._write_recommendations(actions)

        output = self.report_dir / f"tmd_appium_e2e_report_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        workbook.save(output)
        return output.resolve()

    def _write_summary(self, sheet) -> None:
        counts = Counter(result.status for result in self.results)
        total = len(self.results)
        detail_last_row = max(2, total + 1)

        sheet.merge_cells("A1:F1")
        sheet["A1"] = "TMD Care AI Appium E2E Test Report"
        sheet["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        sheet["A1"].fill = PatternFill("solid", fgColor="1F4E79")
        sheet["A1"].alignment = Alignment(horizontal="center")

        meta_rows = [
            ("Run Started", self.run_started_at.strftime("%Y-%m-%d %H:%M:%S")),
            ("Run Finished", self.run_finished_at.strftime("%Y-%m-%d %H:%M:%S")),
            ("Duration (minutes)", round((self.run_finished_at - self.run_started_at).total_seconds() / 60, 2)),
            ("App Package", self.appium_config.app_package),
            ("App Activity", self.appium_config.app_activity),
            ("Device", self.appium_config.device_name),
            ("Appium Server", self.appium_config.server_url),
            ("Raw JSON Results", str(self.raw_results_path.resolve())),
        ]
        for row_index, (label, value) in enumerate(meta_rows, start=3):
            sheet.cell(row=row_index, column=1, value=label)
            sheet.cell(row=row_index, column=2, value=value)

        sheet["A13"] = "Status Summary"
        sheet["A13"].font = Font(bold=True, color="1F4E79")
        sheet["A14"] = "Status"
        sheet["B14"] = "Count"
        sheet["C14"] = "Rate"
        for cell in sheet["14:14"]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4472C4")

        for offset, status in enumerate((STATUS_PASSED, STATUS_FAILED, STATUS_SKIPPED), start=15):
            sheet.cell(row=offset, column=1, value=status)
            sheet.cell(
                row=offset,
                column=2,
                value=f'=COUNTIF(Details!$E$2:$E${detail_last_row},A{offset})',
            )
            sheet.cell(row=offset, column=3, value=f"=IF(SUM($B$15:$B$17)=0,0,B{offset}/SUM($B$15:$B$17))")
            sheet.cell(row=offset, column=3).number_format = "0.0%"

        sheet["A19"] = "Overall Pass Rate"
        sheet["B19"] = "=IF(SUM($B$15:$B$17)=0,0,$B$15/SUM($B$15:$B$17))"
        sheet["B19"].number_format = "0.0%"
        sheet["A20"] = "Collected Tests"
        sheet["B20"] = total
        sheet["A21"] = "Completed Tests"
        sheet["B21"] = counts[STATUS_PASSED] + counts[STATUS_FAILED]

        chart = BarChart()
        chart.title = "Execution Status"
        chart.y_axis.title = "Tests"
        chart.x_axis.title = "Status"
        data = Reference(sheet, min_col=2, min_row=14, max_row=17)
        cats = Reference(sheet, min_col=1, min_row=15, max_row=17)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 7
        chart.width = 12
        sheet.add_chart(chart, "E13")

        self._format_sheet(sheet, widths={"A": 24, "B": 48, "C": 14, "E": 16, "F": 16, "G": 16})

    def _write_details(self, sheet) -> None:
        headers = [
            "Test ID",
            "Suite",
            "Test Name",
            "Markers",
            "Status",
            "Duration (s)",
            "Stage",
            "Failure / Skip Detail",
            "Screenshot",
            "Timestamp",
        ]
        sheet.append(headers)
        for result in self.results:
            sheet.append(
                [
                    result.nodeid,
                    result.suite,
                    result.test_name,
                    result.markers,
                    result.status,
                    result.duration_seconds,
                    result.stage,
                    result.error,
                    "screenshot" if result.screenshot else "",
                    result.timestamp,
                ]
            )
            if result.screenshot:
                cell = sheet.cell(row=sheet.max_row, column=9)
                cell.hyperlink = result.screenshot
                cell.style = "Hyperlink"

        if not self.results:
            sheet.append(["No tests were collected.", "", "", "", STATUS_SKIPPED, 0, "session", "", "", datetime.now().isoformat(timespec="seconds")])

        table_ref = f"A1:J{sheet.max_row}"
        table = Table(displayName="TmdAppiumResults", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        sheet.add_table(table)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = table_ref
        sheet.column_dimensions["A"].width = 54
        sheet.column_dimensions["B"].width = 26
        sheet.column_dimensions["C"].width = 34
        sheet.column_dimensions["D"].width = 36
        sheet.column_dimensions["E"].width = 14
        sheet.column_dimensions["F"].width = 14
        sheet.column_dimensions["G"].width = 12
        sheet.column_dimensions["H"].width = 70
        sheet.column_dimensions["I"].width = 18
        sheet.column_dimensions["J"].width = 22
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=8, max_col=8):
            row[0].alignment = Alignment(wrap_text=True, vertical="top")
        self._apply_status_conditional_formatting(sheet, f"E2:E{sheet.max_row}")

    def _write_coverage(self, sheet) -> None:
        headers = ["Application Area", "Marker", "Representative Coverage", "Requires Credentials", "Observed Status"]
        rows = [
            ("Authentication", "auth", "Login screen, required-field validation, signup, password reset", "No"),
            ("Dashboard Navigation", "navigation", "Dashboard, bottom navigation tabs, core screen reachability", "Yes"),
            ("Pain Tracking", "tracking", "Pain map region selection and save workflow", "Yes"),
            ("Wellness and Sleep", "tracking", "Daily wellness and sleep entry save workflows", "Yes"),
            ("Exercises", "tracking", "Exercise program visibility and start controls", "Yes"),
            ("Progress Reports", "reports", "Progress tabs and health report navigation", "Yes"),
            ("Profile and Support", "support", "Profile, doctors, booking, feedback, and legal/support links", "Yes"),
        ]
        sheet.append(headers)
        for area, marker, coverage, requires_credentials in rows:
            status = self._marker_status(marker)
            sheet.append([area, marker, coverage, requires_credentials, status])

        table_ref = f"A1:E{sheet.max_row}"
        table = Table(displayName="TmdCoverageMatrix", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
        sheet.add_table(table)
        sheet.freeze_panes = "A2"
        sheet.column_dimensions["A"].width = 24
        sheet.column_dimensions["B"].width = 18
        sheet.column_dimensions["C"].width = 70
        sheet.column_dimensions["D"].width = 22
        sheet.column_dimensions["E"].width = 18
        self._apply_status_conditional_formatting(sheet, f"E2:E{sheet.max_row}", include_not_run=True)

    def _write_recommendations(self, sheet) -> None:
        counts = Counter(result.status for result in self.results)
        skipped_credentials = [
            result
            for result in self.results
            if result.status == STATUS_SKIPPED and "requires_credentials" in result.markers
        ]
        recommendations = [
            ("Run environment", "Keep Appium server running and verify the connected phone with adb devices before execution."),
            ("UiAutomator2", "Install or validate the Android driver with appium driver install uiautomator2 and appium driver doctor uiautomator2."),
            ("Stable selectors", "Add Compose test tags or content descriptions to frequently tested controls for more durable Appium locators."),
        ]
        if counts[STATUS_FAILED]:
            recommendations.insert(0, ("Failures", "Open the Details sheet, review failure messages, and inspect linked screenshots."))
        if skipped_credentials:
            recommendations.insert(0, ("Credentials", "Set TMD_TEST_EMAIL and TMD_TEST_PASSWORD to execute the authenticated E2E journeys."))
        if not self.results:
            recommendations.insert(0, ("Collection", "No tests were collected. Run pytest from the appium_tests folder."))

        sheet.append(["Area", "Recommendation"])
        for row in recommendations:
            sheet.append(list(row))
        table_ref = f"A1:B{sheet.max_row}"
        table = Table(displayName="TmdRecommendations", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        sheet.add_table(table)
        sheet.column_dimensions["A"].width = 24
        sheet.column_dimensions["B"].width = 100
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=2):
            row[0].alignment = Alignment(wrap_text=True, vertical="top")

    def _marker_status(self, marker: str) -> str:
        matching = [result for result in self.results if marker in result.markers.split(", ")]
        if not matching:
            return "NOT RUN"
        if any(result.status == STATUS_FAILED for result in matching):
            return STATUS_FAILED
        if any(result.status == STATUS_PASSED for result in matching):
            return STATUS_PASSED
        return STATUS_SKIPPED

    @staticmethod
    def _format_sheet(sheet, widths: dict[str, int] | None = None) -> None:
        sheet.sheet_view.showGridLines = False
        for column, width in (widths or {}).items():
            sheet.column_dimensions[column].width = width
        thin_gray = Side(style="thin", color="D9E2F3")
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = Border(bottom=thin_gray)
                cell.alignment = Alignment(vertical="top")

    @staticmethod
    def _apply_status_conditional_formatting(sheet, range_ref: str, *, include_not_run: bool = False) -> None:
        sheet.conditional_formatting.add(
            range_ref,
            CellIsRule(operator="equal", formula=[f'"{STATUS_PASSED}"'], fill=PatternFill("solid", fgColor="C6EFCE")),
        )
        sheet.conditional_formatting.add(
            range_ref,
            CellIsRule(operator="equal", formula=[f'"{STATUS_FAILED}"'], fill=PatternFill("solid", fgColor="FFC7CE")),
        )
        sheet.conditional_formatting.add(
            range_ref,
            CellIsRule(operator="equal", formula=[f'"{STATUS_SKIPPED}"'], fill=PatternFill("solid", fgColor="FFEB9C")),
        )
        if include_not_run:
            sheet.conditional_formatting.add(
                range_ref,
                CellIsRule(operator="equal", formula=['"NOT RUN"'], fill=PatternFill("solid", fgColor="D9E1F2")),
            )
